import numpy as np
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import os
import scipy as sp
import pandas as pd 
from statsmodels.tsa.tsatools import detrend
import obspy
from scipy.stats import norm
import openpyxl
from openpyxl.styles import Alignment, Font
import base64
import pdfkit

class SummaryResults:
    def __init__(
            self, 
            proyectPath,
            amortiguamientos,
            frecuencias_summary,
            energia_summary,
            Rd_evaluado_lineal_resultado_summary,
            Rd_evaluado_nolineal_resultado_summary,
            num_divisiones_frequ, 
            num_divisiones_ELAS, 
            num_divisiones_INELAS,
            tratamientos
            ):
        print("SummaryResults")
        # Assigning the parameters to instance variables.
        self.proyectPath = proyectPath # Assigning the project path to self.proyectPath
        self.amortiguamientos = amortiguamientos # Assigning the damping ratios to self.amortiguamientos
        self.frecuencias_summary = frecuencias_summary
        self.energia_summary = energia_summary
        self.num_divisiones_frequ = num_divisiones_frequ
        self.num_divisiones_ELAS = num_divisiones_ELAS
        self.num_divisiones_INELAS = num_divisiones_INELAS
        self.tratamientos = tratamientos
       # Define paths for HTML, temporary HTML, and PDF files using the project path and file basename from tratamiento
        self.archivo_html = (
            f"{self.proyectPath}/results/html/summary.html"
        )
        # Define the path for the temporary HTML report file
        self.archivo_temporal_html = f"{self.proyectPath}/results/html/summary.temporal.html"
        # Define the path for the PDF report file
        self.archivo_pdf = (
            f"{self.proyectPath}/results/pdf/summary.pdf"
        ) 
        # Define the path for the Excel report file with suffix '_report_dm'
        self.datos_informe = (
            f"{self.proyectPath}/results/xlsx/summary_report.xlsx"
        )
        # Check if the dynamic amplification report file exists, and if so, delete it
        if os.path.exists(self.archivo_html):
            os.remove(self.archivo_html)

        if os.path.exists(self.archivo_temporal_html):
            os.remove(self.archivo_temporal_html)

        if os.path.exists(self.datos_informe):
            os.remove(self.datos_informe)

        #temporal
        self.directorio_imagenes_general = f"{self.proyectPath}/results/html/images/summary/"
        os.makedirs(self.directorio_imagenes_general, exist_ok=True)

        ###################################################################################################################################
        ############################# Organización de los Resultados de Magnificación dinámica para el rango elastico en una lista de listas
        ###################################################################################################################################
        # Find the maximum length of the sublists that make up Rdlinear containing the calculated dynamic magnification values
        maxfilas_Rdlineal = max(arr.shape[0] for arr in Rd_evaluado_lineal_resultado_summary)

        # Rdlinear Restructure
        reordenar_filas_Rdlineal = [[] for _ in range(maxfilas_Rdlineal)]

        # Rdlinear Restructure
        for arr in Rd_evaluado_lineal_resultado_summary:
            for filas_indexar_Rdlineal in range(arr.shape[0]):
                # Adds the row to the corresponding list in reordenar_filas_Rdlineal
                reordenar_filas_Rdlineal[filas_indexar_Rdlineal].append(arr[filas_indexar_Rdlineal].tolist())

        # Convert grouped rows into list lists called Rdlinear_list
        Rdlineal_list = [row for row in reordenar_filas_Rdlineal]  ############## resultado reordenado


        ###################################################################################################################################
        ############################# Organización de los Resultados de Magnificación dinámica para el rango inelastico en una lista de listas
        ####################################################################################################################################
        # Find the maximum length of the sublists that make up Rdnolinear containing the calculated dynamic magnification values
        maxfilas_Rdnolineal = max(arr.shape[0] for arr in Rd_evaluado_nolineal_resultado_summary)

        # Rdlinear Restructure
        reordenar_filas_Rdnolineal = [[] for _ in range(maxfilas_Rdnolineal)]

        # Rdlinear Restructure
        for arr in Rd_evaluado_nolineal_resultado_summary:
            for filas_indexar_Rdnolineal in range(arr.shape[0]):
                # Adds the row to the corresponding list in reordenar_filas_Rdlineal
                reordenar_filas_Rdnolineal[filas_indexar_Rdnolineal].append(arr[filas_indexar_Rdnolineal].tolist())

        # Convert grouped rows into list lists called Rdlinear_list
        Rdnolineal_list = [row for row in reordenar_filas_Rdnolineal]     ############ resultado reordenado        

        self.Rdlineal_list = Rdlineal_list
        self.Rdnolineal_list = Rdnolineal_list

        self.salidaHtml = []
        self.salidaTemporalHtml = []

        plt.close("all")


    def procesar(self):
        self.localListImages = []
        # Create a directory for the images generated for the summary_report
        if not os.path.exists(self.directorio_imagenes_general):
            os.makedirs(self.directorio_imagenes_general)
        
        etiquetas_x = [f's{i+1}' for i in range(len(self.frecuencias_summary))]
        
        for i in range(len(self.amortiguamientos)):
            
            #######################################################################################################################################
            ########################################## Summary diagram bars of frequencies, energy and dynamic magnification
            #######################################################################################################################################
            # Composite Figure with tree rows and Axis
            fig, (ax1, ax2, ax3) = plt.subplots(nrows=3, figsize=(12,5.80), dpi=300)
            
            # # Calculate additional spacing between groups of bars for frequency and dynamic magnification results.
            sep_grup_barras = 0.7
            ancho_barra = 0.4
            
            # Energy bar graph expressed in percentage and frequencies
            posicion_actual = 0
            posiciones_etiquetas = []
            for grupo_frecuencias, grupo_porcentajes in zip(self.frecuencias_summary, self.energia_summary):
                x_indices = np.arange(len(grupo_frecuencias)) + posicion_actual
                barras = ax1.bar(x_indices, grupo_porcentajes, width=ancho_barra, color='dodgerblue')
            
                # Add the predominant frequencies as labels above each bar
                for barra, frecuencia in zip(barras, grupo_frecuencias):
                    altura = barra.get_height()
                    ax1.text(barra.get_x() + barra.get_width() / 2, altura + 0.4, f'{frecuencia:.2f} [Hz]',
                            ha='center', va='bottom', fontsize=6, rotation=90)
            
                # Obtains the positions of the bars so that the frequencies can be placed above them
                posiciones_etiquetas.append(posicion_actual + (len(grupo_frecuencias) - 1) / 2.0)
            
                # Move the current position to the next group of energy bars
                posicion_actual += len(grupo_frecuencias) + sep_grup_barras
            
            # Configure labels and title for the first graph related to energy frequency and dynamic magnification
            ax1.set_xlabel('')
            ax1.set_ylabel('% Energy', fontsize=7)
            ax1.set_title('')
            ax1.set_xticks(posiciones_etiquetas)  
            ax1.set_xticklabels(etiquetas_x[:len(posiciones_etiquetas)]) 
            por_max = max(max(sublist) for sublist in self.energia_summary)
            ax1.set_ylim(0, por_max * 1.50)
            ax1.tick_params(labelsize=7)
            
            # DM_Elastic Graphic
            posicion_actual = 0
            posiciones_etiquetas = []
            for grupo_dm in self.Rdlineal_list[i]:
                x_indices = np.arange(len(grupo_dm)) + posicion_actual
                barras = ax2.bar(x_indices, grupo_dm, width=ancho_barra, color='skyblue')
            
                # Add the DM_Elastic values as labels above each bar
                for barra, dm in zip(barras, grupo_dm):
                    altura = barra.get_height()
                    ax2.text(barra.get_x() + barra.get_width() / 2, altura + 0.4, f'{dm:.2f}',
                            ha='center', va='bottom', fontsize=6, rotation=90)
            
                # Obtains the positions of the bars so that the DM_Elastic can be placed above them
                posiciones_etiquetas.append(posicion_actual + (len(grupo_dm) - 1) / 2.0)
            
                # Move the current position to the next group of DM_Elastic
                posicion_actual += len(grupo_dm) + sep_grup_barras
            
            # Configure labels and title for the second graph related to elastic dynamic magnification
            ax2.set_xlabel('')
            ax2.set_ylabel('Dynamic Magn. Ψ - Elastic', fontsize=7)
            ax2.set_title('')
            ax2.set_xticks(posiciones_etiquetas)  
            ax2.set_xticklabels(etiquetas_x[:len(posiciones_etiquetas)])  
            dm_ela_max = max(max(sublist) for sublist in self.Rdlineal_list[i])
            ax2.set_ylim(0, dm_ela_max * 1.50)  
            ax2.tick_params(labelsize=7)
            
            # DM_Inelastic graph
            posicion_actual = 0
            posiciones_etiquetas = []
            for grupo_dm in self.Rdnolineal_list[i]:
                x_indices = np.arange(len(grupo_dm)) + posicion_actual
                barras = ax3.bar(x_indices, grupo_dm, width=ancho_barra, color='gray')
            
                # Add the DM_Inelastic values as labels above each bar
                for barra, dm in zip(barras, grupo_dm):
                    altura = barra.get_height()
                    ax3.text(barra.get_x() + barra.get_width() / 2, altura + 0.4, f'{dm:.2f}',
                            ha='center', va='bottom', fontsize=6.0, rotation=90)
            
                # Obtains the positions of the bars so that the DM_Inelastic can be placed above them
                posiciones_etiquetas.append(posicion_actual + (len(grupo_dm) - 1) / 2.0)
            
                # Move the current position to the next group of DM_Inelastic
                posicion_actual += len(grupo_dm) + sep_grup_barras
            
            # Configure labels and title for the second graph related to inelastic dynamic magnification
            ax3.set_xlabel('')
            ax3.set_ylabel('Dynamic Magn. Ψ - Inelastic', fontsize=7.0)
            ax3.set_title('')
            ax3.set_xticks(posiciones_etiquetas)  
            ax3.set_xticklabels(etiquetas_x[:len(posiciones_etiquetas)])  
            dm_inela_max = max(max(sublist) for sublist in self.Rdnolineal_list[i])
            ax3.set_ylim(0, dm_inela_max * 1.50) 
            ax3.tick_params(labelsize=7)
            
            fig.suptitle(f'Bar Diagram Summary of Frequencies, Energy and Dynamic Magnification for ζ={self.amortiguamientos[i]*100:.2f}%', fontsize=8)
            fig.subplots_adjust(top=0.92)
            # Save Bar Diagram Summary of Frequencies, Energy and Dynamic Magnification
            nombre_archivos_graficas_bars = f'summary_Bar_Diagram_Summary_{self.amortiguamientos[i]*100:.2f}_porcentaje.png'  # Nombre del archivo
            self.localListImages.append(nombre_archivos_graficas_bars)
            ruta_completa_imagenes_bars = os.path.join(self.directorio_imagenes_general, nombre_archivos_graficas_bars)
            plt.savefig(ruta_completa_imagenes_bars, bbox_inches='tight')  # Guardar la figura en el directorio
            plt.close(fig)
        
    
        #####################################################################################################################################
        ################################################################# Gaussian Distribution for Frequency
        ####################################################################################################################################
        
            # Flattening frequencies summary
            frecuencias_planas = [f_frequ for sublista_frequ in self.frecuencias_summary for f_frequ in sublista_frequ]
            
            # calculate the minimum and maximum value of frecuencias_planas
            x_min_bins_frequ = min(frecuencias_planas)-0.0001 # tenia problemas con los limites
            x_max_bins_frequ = max(frecuencias_planas)+0.0001
            # self.num_divisiones_frequ = 10
            # Define the limits and frequency intervals through a linspace
            intervalos_frequ = np.linspace(x_min_bins_frequ, x_max_bins_frequ, self.num_divisiones_frequ + 1)
            bins_frequ = intervalos_frequ.tolist()
            
            # create a frequency histogram of the seismic records
            hist_frequ, bin_edges_frequ = np.histogram(frecuencias_planas, bins=bins_frequ, density=True)
            hist_frequ_N, _ = np.histogram(frecuencias_planas, bins=bins_frequ, density=False)
            sum_hist_frequ = np.sum(hist_frequ)
            print(f"La suma de los valores de hist es: {sum_hist_frequ}")
            
            # Deviation Std and Mean for frequencies
            mean_frequ, std_frequ = np.mean(frecuencias_planas), np.std(frecuencias_planas)
            
            # Try to generate a series of values ​​to graph the probability distribution function of the frequencies
            # Try to automatically calculate how many times the minimum and maximum data 
            # differ from the standard deviation from the mean.
            factor_min_frequ = 1 
            umbral_frequ = 0.001
            
            while True:
                x_min_frequ = mean_frequ - factor_min_frequ * std_frequ
                x_max_frequ = mean_frequ + factor_min_frequ * std_frequ
                x_frequ = np.linspace(x_min_frequ, x_max_frequ, 100)
                p_frequ = norm.pdf(x_frequ, mean_frequ, std_frequ)
                
                if p_frequ[0] < umbral_frequ and p_frequ[-1] < umbral_frequ:
                    break
                factor_min_frequ += 1
            
            print('factor seleccionado',factor_min_frequ)
            
            if (mean_frequ - factor_min_frequ * std_frequ) <= min(frecuencias_planas):
                x_min_frequ = mean_frequ - factor_min_frequ * std_frequ
            else:
                x_min_frequ = min(frecuencias_planas)-0.1
            if (mean_frequ + factor_min_frequ * std_frequ) >= max(frecuencias_planas):
                x_max_frequ = mean_frequ + factor_min_frequ * std_frequ
            else:
                x_max_frequ = max(frecuencias_planas)+0.1
                
            # The following steps in order to center the probability distribution function of the frequencies
            if (x_max_frequ - mean_frequ)>=abs(mean_frequ - x_min_frequ):
                distancia_final_frequ = x_max_frequ - mean_frequ
            else:
                distancia_final_frequ = abs(mean_frequ - x_min_frequ)
            
            x_min_frequ = mean_frequ - distancia_final_frequ
            x_max_frequ = mean_frequ + distancia_final_frequ
            # Generate x_frequ values from a minimum to a maximum
            x_frequ = np.linspace(x_min_frequ, x_max_frequ, 100)
            # alculate the values ​​of the probability distribution function for the values ​​of x_frequ
            p_frequ = norm.pdf(x_frequ, mean_frequ, std_frequ)
            
            # Figure in which the probability distribution function of the frequencies and also a histogram will be represented
            fig, ax4 = plt.subplots(figsize=(12, 6.45), dpi=475)
            
            # Plot the grouped histogram of the frequencies of the seismic records
            bars_frequ = ax4.bar(bin_edges_frequ[:-1], hist_frequ, width=np.diff(bin_edges_frequ), alpha=0.6, color='grey', edgecolor='black', align='edge', label='Frequency Histogram')
            
            # Graph the normal distribution of frequencies of the seismic records 
            line_frequ, = ax4.plot(x_frequ, p_frequ, 'k', linewidth=2, label='Gaussian Distribution')
            # Percentile (16,50,84) for Frecuencies 
            percentiles_frequ = np.percentile(frecuencias_planas, [16, 50, 84])
            
            # Shaded areas under the Gauss curve for the 16th 50th 84th percentiles of the frequencies of the seismic records
            x_fill_frequ = np.linspace(x_min_frequ, x_max_frequ, 100)
            x_fill_frequ = np.sort(np.concatenate([x_fill_frequ, percentiles_frequ]))
            p_fill_frequ = norm.pdf(x_fill_frequ, mean_frequ, std_frequ)
            
            
            
            # Probabilities associated with the percentiles (16,50,84) of the frequencies of the seismic records
            p16_frequ = norm.cdf(percentiles_frequ[0], mean_frequ, std_frequ)
            p50_frequ = norm.cdf(percentiles_frequ[1], mean_frequ, std_frequ)
            p84_frequ = norm.cdf(percentiles_frequ[2], mean_frequ, std_frequ)
            
            # Probabilities between percentiles (<=16), (16,50) (50-84) (>=84)
            p_16_frequ = p16_frequ*100
            p_16_50_frequ = (p50_frequ - p16_frequ)*100
            p_50_84_frequ = (p84_frequ - p50_frequ)*100
            p_84_frequ = (1 - p84_frequ)*100
            
            # color the areas related to the 16th, 50th, and 84th percentiles of the frequencies of the seismic records
            ax4.fill_between(x_fill_frequ, p_fill_frequ, where=(x_fill_frequ <= percentiles_frequ[0]), color='SteelBlue', alpha=0.5, label='Area < P16')
            ax4.fill_between(x_fill_frequ, p_fill_frequ, where=((x_fill_frequ >= percentiles_frequ[0]) & (x_fill_frequ <= percentiles_frequ[1])), color='skyblue', alpha=0.5, label='P16 ≤ Area < P50')
            ax4.fill_between(x_fill_frequ, p_fill_frequ, where=((x_fill_frequ >= percentiles_frequ[1]) & (x_fill_frequ <= percentiles_frequ[2])), color='DodgerBlue', alpha=0.5, label='P50 ≤ Area < P84')
            ax4.fill_between(x_fill_frequ, p_fill_frequ, where=(x_fill_frequ >= percentiles_frequ[2]), color='royalblue', alpha=0.5, label='Area ≥ P84')
            
            
            
            
            #etiquetas_frequ = [f"{intervalos_frequ[i]:.2f} - {intervalos_frequ[i+1]:.2f}" for i in range(self.num_divisiones_frequ)]
            
            #for bar_frequ, count_frequ in zip(bars_frequ, hist_frequ_N):
                # Obtener la altura de la barra
                #height_bars_frequ = bar_frequ.get_height()  
            
                # Verificar si `height` es un array y obtener el primer elemento si es necesario
                #if isinstance(height_bars_frequ, (list, np.ndarray)):
                    #height_bars_frequ = height_bars_frequ[0]  # Convertir a un escalar si es necesario
            
                # Verificar si `count` es un array y obtener el primer elemento si es necesario
                #if isinstance(count_frequ, (list, np.ndarray)):
                    #count_frequ = count_frequ[0]  # Convertir a un escalar si es necesario
            
                # Añadir el texto encima de la barra
                #ax4.text(bar_frequ.get_x() + bar_frequ.get_width() / 2, height_bars_frequ + 0.001, f'N={int(count_frequ)}', ha='center', va='bottom', fontsize=7.0, rotation=90)# cambio 6.5 a 7

            # Create labels with the limits of the bars of a histogram of the frequencies of the seismic records
            etiquetas_frequ = [f"({intervalos_frequ[i]:.2f} - {intervalos_frequ[i+1]:.2f})" for i in range(self.num_divisiones_frequ)]
            
            # Place the labels with the limits of the intervals of the grouped frequency histogram in each of its bars    
            for bar_frequ, etiqueta_frequ in zip(bars_frequ, etiquetas_frequ):
                height_bars_frequ = bar_frequ.get_height()  
                
                if isinstance(height_bars_frequ, (list, np.ndarray)):
                    height_bars_frequ = height_bars_frequ[0]  
                
                # Añadir el texto encima de la barra
                ax4.text(bar_frequ.get_x() + bar_frequ.get_width() / 2, 
                        height_bars_frequ + 0.001, 
                        etiqueta_frequ,           
                        ha='center', 
                        va='bottom', 
                        fontsize=7.0, 
                        rotation=90)                   

            
            # Configure the labels, titles, certain limits and parameters of a probability distribution graph of the frequencies of the seismic records accompanied by a histogram
            ax4.set_xlabel('Frequency (Hz)', fontsize=7.5)
            ax4.set_ylabel('Probability Density', fontsize=7.5)
            ax4.set_ylim(0, np.max(hist_frequ) * 1.2)
            ax4.set_title(f'Gaussian Distribution of Frequency for ζ = {self.amortiguamientos[i]*100:.2f}%', pad=18, fontsize=9)
            ax4.set_xlim(x_min_frequ, x_max_frequ)
            ax4.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
            y_max_frequ = np.max(hist_frequ) * 1.2  
            ax4.set_ylim(0, y_max_frequ)
            
            # Select the colors of vertical lines that will be graphed based on the 16,50,84 percentiles
            colores_percentiles_frequ = ['SteelBlue', 'DodgerBlue', 'royalblue']  # Cambia estos colores según prefieras
            
            # for perc_frequ, label_frequ, color_frequ in zip(percentiles_frequ, ['p16', 'p50', 'p84'], colores_percentiles_frequ):
            #     ax.axvline(x=perc_frequ, color=color_frequ, linestyle='--', linewidth=1)
            #     ax.text(perc_frequ-(x_max_frequ-x_min_frequ)*0.01, ax.get_ylim()[1] * 0.83, f'{label_frequ} = {perc_frequ:.2f} Hz', color=color_frequ, ha='center', va='bottom', fontsize=7, rotation=90)
                
            # Plot vertical lines on the 16th, 50th, and 84th percentiles of the frequencies of the seismic records
            for perc_frequ, label_frequ, color_frequ in zip(percentiles_frequ, ['', '', ''], colores_percentiles_frequ):
                ax4.axvline(x=perc_frequ, color=color_frequ, linestyle='--', linewidth=1)
                ax4.text(perc_frequ-(x_max_frequ-x_min_frequ)*0.01, ax4.get_ylim()[1] * 0.83, f'{label_frequ}', color=color_frequ, ha='center', va='bottom', fontsize=7, rotation=90)
            
            # Inicializar los límites inferiores y superiores para DM_ELASTIC y DM_INELASTIC
            dm_elastic_min_max_frequ = {etiqueta_frequ: [np.inf, -np.inf] for etiqueta_frequ in etiquetas_frequ}
            dm_inelastic_min_max_frequ = {etiqueta_frequ: [np.inf, -np.inf] for etiqueta_frequ in etiquetas_frequ}
            frecuencia_min_max_frequ = {etiqueta_frequ: [np.inf, -np.inf] for etiqueta_frequ in etiquetas_frequ}
            
            # Associate the minimum and maximum values ​​of dm_elastic and dm_inelasticto the frequency intervals of the seismic records
            for i_frequ, (freq_sublist_frequ, dm_elastic_sublist_frequ, dm_inelastic_sublist_frequ) in enumerate(zip(self.frecuencias_summary, self.Rdlineal_list[i], self.Rdnolineal_list[i])):  ######################
                for j_frequ, freq_frequ in enumerate(freq_sublist_frequ):
                    for k_frequ in range(self.num_divisiones_frequ):
                        if intervalos_frequ[k_frequ] <= freq_frequ < intervalos_frequ[k_frequ + 1]:
                            frecuencia_min_max_frequ[etiquetas_frequ[k_frequ]][0] = min(frecuencia_min_max_frequ[etiquetas_frequ[k_frequ]][0], freq_frequ)
                            frecuencia_min_max_frequ[etiquetas_frequ[k_frequ]][1] = max(frecuencia_min_max_frequ[etiquetas_frequ[k_frequ]][1], freq_frequ)
                            dm_elastic_min_max_frequ[etiquetas_frequ[k_frequ]][0] = min(dm_elastic_min_max_frequ[etiquetas_frequ[k_frequ]][0], dm_elastic_sublist_frequ[j_frequ])
                            dm_elastic_min_max_frequ[etiquetas_frequ[k_frequ]][1] = max(dm_elastic_min_max_frequ[etiquetas_frequ[k_frequ]][1], dm_elastic_sublist_frequ[j_frequ])
                            dm_inelastic_min_max_frequ[etiquetas_frequ[k_frequ]][0] = min(dm_inelastic_min_max_frequ[etiquetas_frequ[k_frequ]][0], dm_inelastic_sublist_frequ[j_frequ])
                            dm_inelastic_min_max_frequ[etiquetas_frequ[k_frequ]][1] = max(dm_inelastic_min_max_frequ[etiquetas_frequ[k_frequ]][1], dm_inelastic_sublist_frequ[j_frequ])
            
            # Create labels for minimum and maximum values ​​of dm_elastic and dm_inelastic associated with the frequency intervals of the seismic records
            for etiqueta_frequ in etiquetas_frequ:
                if frecuencia_min_max_frequ[etiqueta_frequ][0] == np.inf:
                    frecuencia_min_max_frequ[etiqueta_frequ] = ['N/A', 'N/A']
                if dm_elastic_min_max_frequ[etiqueta_frequ][0] == np.inf:
                    dm_elastic_min_max_frequ[etiqueta_frequ] = ['N/A', 'N/A']
                if dm_inelastic_min_max_frequ[etiqueta_frequ][0] == np.inf:
                    dm_inelastic_min_max_frequ[etiqueta_frequ] = ['N/A', 'N/A']
                if (frecuencia_min_max_frequ[etiqueta_frequ][0] == frecuencia_min_max_frequ[etiqueta_frequ][1] and
                    frecuencia_min_max_frequ[etiqueta_frequ][0] != 'N/A'):
                    frecuencia_min_max_frequ[etiqueta_frequ][1] = frecuencia_min_max_frequ[etiqueta_frequ][0]
                if (dm_elastic_min_max_frequ[etiqueta_frequ][0] == dm_elastic_min_max_frequ[etiqueta_frequ][1] and
                    dm_elastic_min_max_frequ[etiqueta_frequ][0] != 'N/A'):
                    dm_elastic_min_max_frequ[etiqueta_frequ][1] = dm_elastic_min_max_frequ[etiqueta_frequ][0]
                if (dm_inelastic_min_max_frequ[etiqueta_frequ][0] == dm_inelastic_min_max_frequ[etiqueta_frequ][1] and
                    dm_inelastic_min_max_frequ[etiqueta_frequ][0] != 'N/A'):
                    dm_inelastic_min_max_frequ[etiqueta_frequ][1] = dm_inelastic_min_max_frequ[etiqueta_frequ][0]
            
            # Create a group of labels first column frequency intervals second column minimum and maximum values ​​of dm_elastic 
            # and finally third column minimum and maximum values ​​of dm_inelastic
            leyenda_etiquetas_frequ = []
            for etiqueta_frequ in etiquetas_frequ:
                if dm_elastic_min_max_frequ[etiqueta_frequ][0] == 'N/A':
                    elastic_label_frequ = 'N/A'
                else:
                    elastic_label_frequ = f'{dm_elastic_min_max_frequ[etiqueta_frequ][0]:.2f} - {dm_elastic_min_max_frequ[etiqueta_frequ][1]:.2f}'
                
                if dm_inelastic_min_max_frequ[etiqueta_frequ][0] == 'N/A':
                    inelastic_label_frequ = 'N/A'
                else:
                    inelastic_label_frequ = f'{dm_inelastic_min_max_frequ[etiqueta_frequ][0]:.2f} - {dm_inelastic_min_max_frequ[etiqueta_frequ][1]:.2f}'
                
                leyenda_etiquetas_frequ.append(
                    f'{etiqueta_frequ}Hz; ψ_E ({elastic_label_frequ}); ψ_I ({inelastic_label_frequ})'
                )
            
            # Customizing the legends of the graph in which I add the values ​​of the 16th, 50th, 84th percentiles to the Gauss distribution
            # Shade the areas, show the probabilities associated with the already mentioned percentiles and also place the values ​​of frequencies and dynamic magnifications
            handles_frequ = [
                plt.Line2D([0], [0], color=(0, 0, 0, 1), linestyle='-', linewidth=1.5, label='Gaussian Distribution of Frequency'),
                plt.Line2D([0], [0], color='SteelBlue', linestyle='--', linewidth=1, label=f'Percentile 16 = {percentiles_frequ[0]:.2f} Hz'),
                plt.Line2D([0], [0], color='DodgerBlue', linestyle='--', linewidth=1, label=f'Percentile 50 = {percentiles_frequ[1]:.2f} Hz'),
                plt.Line2D([0], [0], color='royalblue', linestyle='--', linewidth=1, label=f'Percentile 86 = {percentiles_frequ[2]:.2f} Hz'),
                plt.Line2D([0], [0], color=(0.27, 0.51, 0.71, 0.5), linestyle='-', linewidth=10, label=f'Pr(Freq≤p16) = {p_16_frequ:.2f}%'), 
                plt.Line2D([0], [0], color=(0.53, 0.81, 0.92, 0.5), linestyle='-', linewidth=10, label=f'Pr(p16<Freq≤p50) = {p_16_50_frequ:.2f}%'),
                plt.Line2D([0], [0], color=(0.12, 0.56, 1, 0.5), linestyle='-', linewidth=10, label=f'Pr(p50<Freq≤p84) = {p_50_84_frequ:.2f}%'),
                plt.Line2D([0], [0], color=(0.25, 0.41, 0.88, 0.5), linestyle='-', linewidth=10, label=f'Pr(Freq>p84) = {p_84_frequ:.2f}%'),
                plt.Line2D([0], [0], color=(0.5, 0.5, 0.5, 0.5), linestyle='-', linewidth=10, label='Frequency Histogram'),
                plt.Line2D([0], [0], color='none', linestyle='', linewidth=0, label='Frequency Intervals and Dynamic Magnification:')
            ]
            handles_frequ.extend([
                plt.Line2D([0], [0], color='white', linestyle='-', linewidth=0, marker='o', markersize=1, label=etiqueta_frequ)
                for etiqueta_frequ in leyenda_etiquetas_frequ
            ])
            
            # Addsymbols of the mean and standard deviation followed by their values ​​for the frequency distribution of the seismic records
            mean_text_frequ = f'$\overline{{x}} = {mean_frequ:.2f} \ \mathrm{{Hz}}$'
            std_text_frequ = f'σ = {std_frequ:.2f} Hz'
            ax4.text(0.01, 0.98, mean_text_frequ, transform=ax4.transAxes, fontsize=7.5, verticalalignment='top', horizontalalignment='left')
            ax4.text(0.01, 0.93, std_text_frequ, transform=ax4.transAxes, fontsize=7.5, verticalalignment='top', horizontalalignment='left')
            # parameters to customize the graph adjust legends text sizes
            ax4.legend(handles=handles_frequ, loc='upper left', bbox_to_anchor=(1, 1), title=None, frameon=False, fontsize=7.0)
            ax4.yaxis.set_ticks([])
            plt.xticks(fontsize=7.5)  
            plt.yticks(fontsize=7.5)  
            plt.tight_layout()
            # Save plot summary Gaussian frequency distribution of seismic records
            nombre_archivos_graficas_frecuencias_distribuciong = f'symmary_Gaussian_Distribution_of_Frequency_for_zeta_{self.amortiguamientos[i]*100:.2f}_porcentaje.png'
            self.localListImages.append(nombre_archivos_graficas_frecuencias_distribuciong)
            ruta_completa_imagenes_freq_distribuciong = os.path.join(self.directorio_imagenes_general, nombre_archivos_graficas_frecuencias_distribuciong)
            plt.savefig(ruta_completa_imagenes_freq_distribuciong, bbox_inches='tight')  # Guardar la figura en el directorio
            plt.close(fig)

            #############################################################################################################################
            ####################################################################### DISTRIBUCION GAUSIANA DE LA MAGNIFICACION DINAMICA ELASTICA
            ##########################################################################################################################
            
            # Flattening Rdlineal_list
            DM_ELASTIC_PLANO = [f_ELAS for sublista_ELAS in self.Rdlineal_list[i] for f_ELAS in sublista_ELAS]
            
            # # calculate the minimum and maximum value of Rdlineal_list
            x_min_bins_ELAS = min(DM_ELASTIC_PLANO)-0.0001
            x_max_bins_ELAS = max(DM_ELASTIC_PLANO)+0.0001
            # self.num_divisiones_ELAS = 7
            # Define the limits and dynamic magnification elastic intervals through a linspace
            intervalos_ELAS = np.linspace(x_min_bins_ELAS, x_max_bins_ELAS, self.num_divisiones_ELAS + 1)
            bins_ELAS = intervalos_ELAS.tolist()
            
            
            # Create a elastic dynamic magnification histogram of the seismic records
            hist_ELAS, bin_edges_ELAS = np.histogram(DM_ELASTIC_PLANO, bins=bins_ELAS, density=True)
            hist_ELAS_N, _ = np.histogram(DM_ELASTIC_PLANO, bins=bins_ELAS, density=False)
            
            sum_hist_ELAS = np.sum(hist_ELAS)
            print(f"La suma de los valores de hist es: {sum_hist_ELAS}")
            
            # Deviation Std and Mean for elastic dynamic magnification
            mean_ELAS, std_ELAS = np.mean(DM_ELASTIC_PLANO), np.std(DM_ELASTIC_PLANO)
            
            # Try to generate a series of values ​​to graph the probability distribution function of the elastic dynamic magnification
            # Try to automatically calculate how many times the minimum and maximum data 
            # differ from the standard deviation from the mean.
            factor_min_ELAS = 1
            
            umbral_ELAS = 0.001
            
            
            while True:
                x_min_ELAS = mean_ELAS - factor_min_ELAS * std_ELAS
                x_max_ELAS = mean_ELAS + factor_min_ELAS * std_ELAS
                x_ELAS = np.linspace(x_min_ELAS, x_max_ELAS, 100)
                p_ELAS = norm.pdf(x_ELAS, mean_ELAS, std_ELAS)
                
                if p_ELAS[0] < umbral_ELAS and p_ELAS[-1] < umbral_ELAS:
                    break
                factor_min_ELAS += 1
            
            print('promedio dyanmic elastico',mean_ELAS)
            print('desviacion_estandar elastico',std_ELAS)
            print('factor seleccionado dyanmic elastico',factor_min_ELAS)
            
            if (mean_ELAS - factor_min_ELAS * std_ELAS) <= min(DM_ELASTIC_PLANO):
                x_min_ELAS = mean_ELAS - factor_min_ELAS * std_ELAS
            else:
                x_min_ELAS = min(DM_ELASTIC_PLANO)-0.1
            if (mean_ELAS + factor_min_ELAS * std_ELAS) >= max(DM_ELASTIC_PLANO):
                x_max_ELAS = mean_ELAS + factor_min_ELAS * std_ELAS
            else:
                x_max_ELAS = max(DM_ELASTIC_PLANO)+0.1
                
            
            # The following steps in order to center the probability distribution function of the elastic dynamic magnification
            if (x_max_ELAS - mean_ELAS)>=abs(mean_ELAS - x_min_ELAS):
                distancia_final_ELAS = x_max_ELAS - mean_ELAS
            else:
                distancia_final_ELAS = abs(mean_ELAS - x_min_ELAS)
            
            x_min_ELAS = mean_ELAS - distancia_final_ELAS
            x_max_ELAS = mean_ELAS + distancia_final_ELAS
            
            # Generate x_ELAS values from a minimum to a maximum
            x_ELAS = np.linspace(x_min_ELAS, x_max_ELAS, 100)
            # alculate the values ​​of the probability distribution function for the values ​​of x_ELAS
            p_ELAS = norm.pdf(x_ELAS, mean_ELAS, std_ELAS)
            
            # Figure in which the probability distribution function of the elastic dynamic magnification and also a histogram will be represented
            fig, ax5 = plt.subplots(figsize=(12, 6.45), dpi=475)
            
            # Plot the grouped histogram of the elastic dynamic magnification 
            bars_ELAS = ax5.bar(bin_edges_ELAS[:-1], hist_ELAS, width=np.diff(bin_edges_ELAS), alpha=0.6, color='grey', edgecolor='black', align='edge', label='E. Dynamic Magnification Histogram')
            
            # Graph the normal distribution of elastic dynamic magnification 
            line_ELAS, = ax5.plot(x_ELAS, p_ELAS, 'k', linewidth=2, label='Gaussian Distribution')
            # Percentile (16,50,84) for Elastic Dynamic Magnification
            percentiles_ELAS = np.percentile(DM_ELASTIC_PLANO, [16, 50, 84])
            
            # Shaded areas under the Gauss curve for the 16th 50th 84th percentiles of the elastic dynamic magnification 
            x_fill_ELAS = np.linspace(x_min_ELAS, x_max_ELAS, 100)
            x_fill_ELAS = np.sort(np.concatenate([x_fill_ELAS, percentiles_ELAS]))
            p_fill_ELAS = norm.pdf(x_fill_ELAS, mean_ELAS, std_ELAS)
            
            
            
            # Probabilities associated with the percentiles (16,50,84) of the elastic dynamic magnification 
            p16_ELAS = norm.cdf(percentiles_ELAS[0], mean_ELAS, std_ELAS)
            p50_ELAS = norm.cdf(percentiles_ELAS[1], mean_ELAS, std_ELAS)
            p84_ELAS = norm.cdf(percentiles_ELAS[2], mean_ELAS, std_ELAS)
            
            # Probabilities between percentiles (<=16), (16,50) (50-84) (>=84)
            p_16_ELAS = p16_ELAS*100
            p_16_50_ELAS = (p50_ELAS - p16_ELAS)*100
            p_50_84_ELAS = (p84_ELAS - p50_ELAS)*100
            p_84_ELAS = (1 - p84_ELAS)*100
            
            
            # color the areas related to the 16th, 50th, and 84th percentiles of the elastic dynamic magnification 
            ax5.fill_between(x_fill_ELAS, p_fill_ELAS, where=(x_fill_ELAS <= percentiles_ELAS[0]), color='SteelBlue', alpha=0.5, label='Area < P16')
            ax5.fill_between(x_fill_ELAS, p_fill_ELAS, where=((x_fill_ELAS >= percentiles_ELAS[0]) & (x_fill_ELAS <= percentiles_ELAS[1])), color='skyblue', alpha=0.5, label='P16 ≤ Area < P50')
            ax5.fill_between(x_fill_ELAS, p_fill_ELAS, where=((x_fill_ELAS >= percentiles_ELAS[1]) & (x_fill_ELAS <= percentiles_ELAS[2])), color='DodgerBlue', alpha=0.5, label='P50 ≤ Area < P84')
            ax5.fill_between(x_fill_ELAS, p_fill_ELAS, where=(x_fill_ELAS >= percentiles_ELAS[2]), color='royalblue', alpha=0.5, label='Area ≥ P84')
            
            #etiquetas_ELAS = [f"{intervalos_ELAS[i]:.2f} - {intervalos_ELAS[i+1]:.2f}" for i in range(self.num_divisiones_ELAS)]
            
            #for bar_ELAS, count_ELAS in zip(bars_ELAS, hist_ELAS_N):
                # Obtener la altura de la barra
                #height_bars_ELAS = bar_ELAS.get_height()  
            
                # Verificar si `height` es un array y obtener el primer elemento si es necesario
                #if isinstance(height_bars_ELAS, (list, np.ndarray)):
                    #height_bars_ELAS = height_bars_ELAS[0]  # Convertir a un escalar si es necesario
            
                # Verificar si `count` es un array y obtener el primer elemento si es necesario
                #if isinstance(count_ELAS, (list, np.ndarray)):
                    #count_ELAS = count_ELAS[0]  # Convertir a un escalar si es necesario
            
                # Añadir el texto encima de la barra
                #ax5.text(bar_ELAS.get_x() + bar_ELAS.get_width() / 2, height_bars_ELAS + 0.01, f'N={int(count_ELAS)}', ha='center', va='bottom', fontsize=7.0, rotation=90)
            
            # Create labels with the limits of the bars of a histogram of the dynamic magnification

            etiquetas_ELAS = [f"({intervalos_ELAS[i]:.2f} - {intervalos_ELAS[i+1]:.2f})" for i in range(self.num_divisiones_ELAS)]
            
            # Place the labels with the limits of the intervals of the grouped elastic dynamic magnification histogram in each of its bars     
            
            for bar_ELAS, etiqueta_ELAS in zip(bars_ELAS, etiquetas_ELAS):
                height_bars_ELAS = bar_ELAS.get_height()  
                if isinstance(height_bars_ELAS, (list, np.ndarray)):
                    height_bars_ELAS = height_bars_ELAS[0]  
                
                ax5.text(bar_ELAS.get_x() + bar_ELAS.get_width() / 2, 
                        height_bars_ELAS + 0.002, 
                        etiqueta_ELAS,                
                        ha='center', 
                        va='bottom', 
                        fontsize=7.0, 
                        rotation=90)                   
            
            
            # Configure the labels, titles, certain limits and parameters of a probability distribution graph of the elastic dynamic magnification accompanied by a histogram
            ax5.set_xlabel('ψ Elastic Dynamic Magnification', fontsize=7.5)
            ax5.set_ylabel('Probability Density', fontsize=7.5)
            ax5.set_ylim(0, np.max(hist_ELAS) * 1.2)
            ax5.set_title(f'Gaussian Distribution of Elastic Dynamic Magnification for ζ = {self.amortiguamientos[i]*100:.2f}%', pad=18, fontsize=9)
            ax5.set_xlim(x_min_ELAS, x_max_ELAS)
            ax5.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
            y_max_ELAS = np.max(hist_ELAS) * 1.2 
            ax5.set_ylim(0, y_max_ELAS)
            
            # Select the colors of vertical lines that will be graphed based on the 16,50,84 percentiles
            colores_percentiles_ELAS = ['SteelBlue', 'DodgerBlue', 'royalblue']  # Cambia estos colores según prefieras
            
            # # Agregar líneas verticales para los percentiles 16, 50 y 84 con etiquetas
            # for perc_ELAS, label_ELAS, color_ELAS in zip(percentiles_ELAS, ['p16', 'p50', 'p84'], colores_percentiles_ELAS):
            #     ax.axvline(x=perc_ELAS, color=color_ELAS, linestyle='--', linewidth=1)
            #     ax.text(perc_ELAS-(x_max_ELAS-x_min_ELAS)*0.01, ax.get_ylim()[1] * 0.83, f'{label_ELAS} = {perc_ELAS:.2f} Hz', color=color_ELAS, ha='center', va='bottom', fontsize=7, rotation=90)
              
            # Plot vertical lines on the 16th, 50th, and 84th percentiles of the elastic dynamic magnification   
            for perc_ELAS, label_ELAS, color_ELAS in zip(percentiles_ELAS, ['', '', ''], colores_percentiles_ELAS):
                ax5.axvline(x=perc_ELAS, color=color_ELAS, linestyle='--', linewidth=1)
                ax5.text(perc_ELAS-(x_max_ELAS-x_min_ELAS)*0.01, ax5.get_ylim()[1] * 0.83, f'{label_ELAS}', color=color_ELAS, ha='center', va='bottom', fontsize=7, rotation=90)
            
            dm_elastic_min_max_ELAS = {etiqueta_ELAS: [np.inf, -np.inf] for etiqueta_ELAS in etiquetas_ELAS}
            frecuencia_min_max_ELAS = {etiqueta_ELAS: [np.inf, -np.inf] for etiqueta_ELAS in etiquetas_ELAS}
            
            # Associate the minimum and maximum values ​​of dm_inelastic and frequencies the elastic dynamic magnification intervals
            for i_ELAS, (ELAS_SUBLIST_ELAS, frequ_elastic_sublist_ELAS) in enumerate(zip(self.Rdlineal_list[i], self.frecuencias_summary)):
                for j_ELAS, freq_ELAS in enumerate(ELAS_SUBLIST_ELAS):
                    for k_ELAS in range(self.num_divisiones_ELAS):
                        if intervalos_ELAS[k_ELAS] <= freq_ELAS < intervalos_ELAS[k_ELAS + 1]:
                            frecuencia_min_max_ELAS[etiquetas_ELAS[k_ELAS]][0] = min(frecuencia_min_max_ELAS[etiquetas_ELAS[k_ELAS]][0], freq_ELAS)
                            frecuencia_min_max_ELAS[etiquetas_ELAS[k_ELAS]][1] = max(frecuencia_min_max_ELAS[etiquetas_ELAS[k_ELAS]][1], freq_ELAS)
                            dm_elastic_min_max_ELAS[etiquetas_ELAS[k_ELAS]][0] = min(dm_elastic_min_max_ELAS[etiquetas_ELAS[k_ELAS]][0], frequ_elastic_sublist_ELAS[j_ELAS])
                            dm_elastic_min_max_ELAS[etiquetas_ELAS[k_ELAS]][1] = max(dm_elastic_min_max_ELAS[etiquetas_ELAS[k_ELAS]][1], frequ_elastic_sublist_ELAS[j_ELAS])
            
            # Create labels for minimum and maximum values ​​of dm_inelastic and frequency associated with the elastic dynamic magnification
            for etiqueta_ELAS in etiquetas_ELAS:
                if frecuencia_min_max_ELAS[etiqueta_ELAS][0] == np.inf:
                    frecuencia_min_max_ELAS[etiqueta_ELAS] = ['N/A', 'N/A']
                if dm_elastic_min_max_ELAS[etiqueta_ELAS][0] == np.inf:
                    dm_elastic_min_max_ELAS[etiqueta_ELAS] = ['N/A', 'N/A']
                if (frecuencia_min_max_ELAS[etiqueta_ELAS][0] == frecuencia_min_max_ELAS[etiqueta_ELAS][1] and
                    frecuencia_min_max_ELAS[etiqueta_ELAS][0] != 'N/A'):
                    frecuencia_min_max_ELAS[etiqueta_ELAS][1] = frecuencia_min_max_ELAS[etiqueta_ELAS][0]
                if (dm_elastic_min_max_ELAS[etiqueta_ELAS][0] == dm_elastic_min_max_ELAS[etiqueta_ELAS][1] and
                    dm_elastic_min_max_ELAS[etiqueta_ELAS][0] != 'N/A'):
                    dm_elastic_min_max_ELAS[etiqueta_ELAS][1] = dm_elastic_min_max_ELAS[etiqueta_ELAS][0]
            
            # Create a group of labels first column elastic dynamic magnification intervals second column minimum and maximum values ​​of dm_inelastic
            # and finally third column minimum and maximum values ​​of frequencies
            leyenda_etiquetas_ELAS = []
            for etiqueta_ELAS in etiquetas_ELAS:
                if dm_elastic_min_max_ELAS[etiqueta_ELAS][0] == 'N/A':
                    elastic_label_ELAS = 'N/A'
                else:
                    elastic_label_ELAS = f'{dm_elastic_min_max_ELAS[etiqueta_ELAS][0]:.2f} - {dm_elastic_min_max_ELAS[etiqueta_ELAS][1]:.2f}'
                
                leyenda_etiquetas_ELAS.append(
                    f'ψ_E {etiqueta_ELAS}; Freq. ({elastic_label_ELAS})Hz'
                )
            
            # Customizing the legends of the graph in which I add the values ​​of the 16th, 50th, 84th percentiles to the Gauss distribution
            # Shade the areas, show the probabilities associated with the already mentioned percentiles and also place the values ​​of frequencies and dynamic magnifications
            handles_ELAS = [
                plt.Line2D([0], [0], color=(0, 0, 0, 1), linestyle='-', linewidth=2, label='Gaussian Distribution of Elastic Dynamic Magnification'),
                plt.Line2D([0], [0], color='SteelBlue', linestyle='--', linewidth=1, label=f'Percentile 16 = {percentiles_ELAS[0]:.2f}'),
                plt.Line2D([0], [0], color='DodgerBlue', linestyle='--', linewidth=1, label=f'Percentile 50 = {percentiles_ELAS[1]:.2f}'),
                plt.Line2D([0], [0], color='royalblue', linestyle='--', linewidth=1, label=f'Percentile 86 = {percentiles_ELAS[2]:.2f}'),
                plt.Line2D([0], [0], color=(0.27, 0.51, 0.71, 0.5), linestyle='-', linewidth=10, label=f'Pr(ψ_E≤p16) = {p_16_ELAS:.2f}%'),
                plt.Line2D([0], [0], color=(0.53, 0.81, 0.92, 0.5), linestyle='-', linewidth=10, label=f'Pr(p16<ψ_E≤p50) = {p_16_50_ELAS:.2f}%'),
                plt.Line2D([0], [0], color=(0.12, 0.56, 1, 0.5), linestyle='-', linewidth=10, label=f'Pr(p50<ψ_E≤p84) = {p_50_84_ELAS:.2f}%'),
                plt.Line2D([0], [0], color=(0.25, 0.41, 0.88, 0.5), linestyle='-', linewidth=10, label=f'Pr(ψ_E>p84) = {p_84_ELAS:.2f}%'),
                plt.Line2D([0], [0], color=(0.5, 0.5, 0.5, 0.5), linestyle='-', linewidth=10, label='E. Dynamic Magnification Histogram')
                #plt.Line2D([0], [0], color='none', linestyle='', linewidth=0, label='Elastic Dynamic Magnification Intervals and Frequency:')
            ]
            
            #handles_ELAS.extend([
                #plt.Line2D([0], [0], color='white', linestyle='-', linewidth=0, marker='o', markersize=1, label=etiqueta_ELAS)
                #for etiqueta_ELAS in leyenda_etiquetas_ELAS
            #])
            
            # Addsymbols of the mean and standard deviation followed by their values ​​for the elastic dynamic magnification
            mean_text_ELAS = f'$\overline{{x}} = {mean_ELAS:.2f} \ \mathrm{{Hz}}$'
            std_text_ELAS = f'σ = {std_ELAS:.2f} Hz'
            ax5.text(0.01, 0.98, mean_text_ELAS, transform=ax5.transAxes, fontsize=7.5, verticalalignment='top', horizontalalignment='left')
            ax5.text(0.01, 0.93, std_text_ELAS, transform=ax5.transAxes, fontsize=7.5, verticalalignment='top', horizontalalignment='left')
            
            # parameters to customize the graph adjust legends text sizes
            ax5.legend(handles=handles_ELAS, loc='upper left', bbox_to_anchor=(1, 1), title=None, frameon=False, fontsize=7.0) 
            ax5.yaxis.set_ticks([])
            plt.xticks(fontsize=7.5)  
            plt.yticks(fontsize=7.5)  
            plt.tight_layout()
            # Save plot summary Gaussian Distribution of Elastic Dynamic Magnification
            nombre_archivos_graficas_MDIN_ELAS_distribuciong = f'summary_Gaussian_Distribution_of_Elastic_Dynamic_Magnification_for_zeta_{self.amortiguamientos[i]*100:.2f}_porcentaje.png'
            self.localListImages.append(nombre_archivos_graficas_MDIN_ELAS_distribuciong)
            ruta_completa_imagenes_MDIN_ELAS_distribuciong = os.path.join(self.directorio_imagenes_general, nombre_archivos_graficas_MDIN_ELAS_distribuciong)
            plt.savefig(ruta_completa_imagenes_MDIN_ELAS_distribuciong, bbox_inches='tight')  # Guardar la figura en el directorio
            plt.close(fig)
            #####################################################################################################################################
            ###################################################################### Distribución de Gauss de la Magnificación Dinámica Inelastica
            #####################################################################################################################################

            # # Flattening Rdnolineal_list
            DM_INELASTIC_PLANO = [f_INELAS for sublista_INELAS in self.Rdnolineal_list[i] for f_INELAS in sublista_INELAS]
            
            # Calculate the minimum and maximum value of Rdlineal_list
            x_min_bins_INELAS = min(DM_INELASTIC_PLANO)-0.0001
            x_max_bins_INELAS = max(DM_INELASTIC_PLANO)+0.0001
            # self.num_divisiones_INELAS = 7
            # Define the limits and dynamic magnification inelastic intervals through a linspace
            intervalos_INELAS = np.linspace(x_min_bins_INELAS, x_max_bins_INELAS, self.num_divisiones_INELAS + 1)
            bins_INELAS = intervalos_INELAS.tolist()
            
            
            # Create a inelastic dynamic magnification histogram of the seismic records
            hist_INELAS, bin_edges_INELAS = np.histogram(DM_INELASTIC_PLANO, bins=bins_INELAS, density=True)
            hist_INELAS_N, _ = np.histogram(DM_INELASTIC_PLANO, bins=bins_INELAS, density=False)
            
            sum_hist_INELAS = np.sum(hist_INELAS)
            print(f"La suma de los valores de hist es: {sum_hist_INELAS}")
            
            # Deviation Std and Mean for inelastic dynamic magnification
            mean_INELAS, std_INELAS = np.mean(DM_INELASTIC_PLANO), np.std(DM_INELASTIC_PLANO)
            
            # Try to generate a series of values ​​to graph the probability distribution function of the inelastic dynamic magnification
            # Try to automatically calculate how many times the minimum and maximum data 
            # differ from the standard deviation from the mean.
            
            factor_min_INELAS = 1
            
            umbral_INELAS = 0.001
            
            while True:
                x_min_INELAS = mean_INELAS - factor_min_INELAS * std_INELAS
                x_max_INELAS = mean_INELAS + factor_min_INELAS * std_INELAS
                x_INELAS = np.linspace(x_min_INELAS, x_max_INELAS, 100)
                p_INELAS = norm.pdf(x_INELAS, mean_INELAS, std_INELAS)
                
                if p_INELAS[0] < umbral_INELAS and p_INELAS[-1] < umbral_INELAS:
                    break
                # Aumentar los factores si no se cumple la condición
                factor_min_INELAS += 1
            
            print('promedio dyanmic elastico',mean_INELAS)
            print('desviacion_estandar elastico',std_INELAS)
            print('factor seleccionado dyanmic elastico',factor_min_INELAS)
            
            if (mean_INELAS - factor_min_INELAS * std_INELAS) <= min(DM_INELASTIC_PLANO):
                x_min_INELAS = mean_INELAS - factor_min_INELAS * std_INELAS
            else:
                x_min_INELAS = min(DM_INELASTIC_PLANO)-0.1
            if (mean_INELAS + factor_min_INELAS * std_INELAS) >= max(DM_INELASTIC_PLANO):
                x_max_INELAS = mean_INELAS + factor_min_INELAS * std_INELAS
            else:
                x_max_INELAS = max(DM_INELASTIC_PLANO)+0.1
                
            
             # The following steps in order to center the probability distribution function of the inelastic dynamic magnification
            if (x_max_INELAS - mean_INELAS)>=abs(mean_INELAS - x_min_INELAS):
                distancia_final_INELAS = x_max_INELAS - mean_INELAS
            else:
                distancia_final_INELAS = abs(mean_INELAS - x_min_INELAS)
            
            print ()
            x_min_INELAS = mean_INELAS - distancia_final_INELAS
            x_max_INELAS = mean_INELAS + distancia_final_INELAS
            
            
            x_INELAS = np.linspace(x_min_INELAS, x_max_INELAS, 100)
            p_INELAS = norm.pdf(x_INELAS, mean_INELAS, std_INELAS)
            
            # Figure in which the probability distribution function of the inelastic dynamic magnification and also a histogram will be represented
            fig, ax6 = plt.subplots(figsize=(12, 6.45), dpi=475)
            
            # Plot the grouped histogram of the inelastic dynamic magnification
            bars_INELAS = ax6.bar(bin_edges_INELAS[:-1], hist_INELAS, width=np.diff(bin_edges_INELAS), alpha=0.6, color='grey', edgecolor='black', align='edge', label='I. Dynamic Magnification Histogram')
            
            # Graph the normal distribution of inelastic dynamic magnification
            line_INELAS, = ax6.plot(x_INELAS, p_INELAS, 'k', linewidth=2, label='Gaussian Distribution of Inelastic Dynamic Magnification')
            # Percentile (16,50,84) for Inelastic Dynamic Magnification
            percentiles_INELAS = np.percentile(DM_INELASTIC_PLANO, [16, 50, 84])
            
            # Shaded areas under the Gauss curve for the 16th 50th 84th percentiles of the inelastic dynamic magnification 
            x_fill_INELAS = np.linspace(x_min_INELAS, x_max_INELAS, 100)
            x_fill_INELAS = np.sort(np.concatenate([x_fill_INELAS, percentiles_INELAS]))
            p_fill_INELAS = norm.pdf(x_fill_INELAS, mean_INELAS, std_INELAS)
            
            
            
            # Probabilities associated with the percentiles (16,50,84) of the inelastic dynamic magnification 
            p16_INELAS = norm.cdf(percentiles_INELAS[0], mean_INELAS, std_INELAS)
            p50_INELAS = norm.cdf(percentiles_INELAS[1], mean_INELAS, std_INELAS)
            p84_INELAS = norm.cdf(percentiles_INELAS[2], mean_INELAS, std_INELAS)
            
            # Probabilities between percentiles (<=16), (16,50) (50-84) (>=84)
            p_16_INELAS = p16_INELAS*100
            p_16_50_INELAS = (p50_INELAS - p16_INELAS)*100
            p_50_84_INELAS = (p84_INELAS - p50_INELAS)*100
            p_84_INELAS = (1 - p84_INELAS)*100
            
            
            # Color the areas related to the 16th, 50th, and 84th percentiles of the inelastic dynamic magnification
            ax6.fill_between(x_fill_INELAS, p_fill_INELAS, where=(x_fill_INELAS <= percentiles_INELAS[0]), color='SteelBlue', alpha=0.5, label='Area < P16')
            ax6.fill_between(x_fill_INELAS, p_fill_INELAS, where=((x_fill_INELAS >= percentiles_INELAS[0]) & (x_fill_INELAS <= percentiles_INELAS[1])), color='skyblue', alpha=0.5, label='P16 ≤ Area < P50')
            ax6.fill_between(x_fill_INELAS, p_fill_INELAS, where=((x_fill_INELAS >= percentiles_INELAS[1]) & (x_fill_INELAS <= percentiles_INELAS[2])), color='DodgerBlue', alpha=0.5, label='P50 ≤ Area < P84')
            ax6.fill_between(x_fill_INELAS, p_fill_INELAS, where=(x_fill_INELAS >= percentiles_INELAS[2]), color='royalblue', alpha=0.5, label='Area ≥ P84')
            
            # Etiquetas en las barras del histograma
            #for bar_INELAS, count_INELAS in zip(bars_INELAS, hist_INELAS_N):
                # Obtener la altura de la barra
                #height_bars_INELAS = bar_INELAS.get_height()  
            
                # Verificar si `height` es un array y obtener el primer elemento si es necesario
                #if isinstance(height_bars_INELAS, (list, np.ndarray)):
                    #height_bars_INELAS = height_bars_INELAS[0]  # Convertir a un escalar si es necesario
            
                # Verificar si `count` es un array y obtener el primer elemento si es necesario
                #if isinstance(count_INELAS, (list, np.ndarray)):
                    #count_INELAS = count_INELAS[0]  # Convertir a un escalar si es necesario
            
                # Añadir el texto encima de la barra
                #ax6.text(bar_INELAS.get_x() + bar_INELAS.get_width() / 2, height_bars_INELAS + 0.01, f'N={int(count_INELAS)}', ha='center', va='bottom', fontsize=7.0, rotation=90)
            
            # Create labels with the limits of the bars of a histogram of the dynamic magnification
            etiquetas_INELAS = [f"({intervalos_INELAS[i]:.2f} - {intervalos_INELAS[i+1]:.2f})" for i in range(self.num_divisiones_INELAS)]
             
            # Place the labels with the limits of the intervals of the grouped inelastic dynamic magnification histogram in each of its bars       
            for bar_INELAS, etiqueta_INELAS in zip(bars_INELAS, etiquetas_INELAS):
                height_bars_INELAS = bar_INELAS.get_height()  
                
                if isinstance(height_bars_INELAS, (list, np.ndarray)):
                    height_bars_INELAS = height_bars_INELAS[0]  
                
                # Añadir el texto encima de la barra
                ax6.text(bar_INELAS.get_x() + bar_INELAS.get_width() / 2, 
                        height_bars_INELAS + 0.002, 
                        etiqueta_INELAS,                
                        ha='center', 
                        va='bottom', 
                        fontsize=7.0, 
                        rotation=90)                   
            
            
            # Configure the labels, titles, certain limits and parameters of a probability distribution graph of the inelastic dynamic magnification accompanied by a histogram
            ax6.set_xlabel('ψ  Inelastic Dynamic Magnification', fontsize=7.5)
            ax6.set_ylabel('Probability Density', fontsize=7.5)
            ax6.set_ylim(0, np.max(hist_INELAS) * 1.2)
            ax6.set_title(f'Gaussian Distribution of Inelastic Dynamic Magnification for ζ = {self.amortiguamientos[i]*100:.2f}%', pad=18, fontsize=9)
            ax6.set_xlim(x_min_INELAS, x_max_INELAS)
            ax6.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
            y_max_INELAS = np.max(hist_INELAS) * 1.2  
            ax6.set_ylim(0, y_max_INELAS)
            
            
            # Select the colors of vertical lines that will be graphed based on the 16,50,84 percentiles
            colores_percentiles_INELAS = ['SteelBlue', 'DodgerBlue', 'royalblue'] 
            
            # # Agregar líneas verticales para los percentiles 16, 50 y 84 con etiquetas
            # for perc_INELAS, label_INELAS, color_INELAS in zip(percentiles_INELAS, ['p16', 'p50', 'p84'], colores_percentiles_INELAS):
            #     ax.axvline(x=perc_INELAS, color=color_INELAS, linestyle='--', linewidth=1)
            #     ax.text(perc_INELAS-(x_max_INELAS-x_min_INELAS)*0.01, ax.get_ylim()[1] * 0.83, f'{label_INELAS} = {perc_INELAS:.2f} Hz', color=color_INELAS, ha='center', va='bottom', fontsize=7, rotation=90)
                
            # Plot vertical lines on the 16th, 50th, and 84th percentiles of the inelastic dynamic magnification
            for perc_INELAS, label_INELAS, color_INELAS in zip(percentiles_INELAS, ['', '', ''], colores_percentiles_INELAS):
                ax6.axvline(x=perc_INELAS, color=color_INELAS, linestyle='--', linewidth=1)
                ax6.text(perc_INELAS-(x_max_INELAS-x_min_INELAS)*0.01, ax6.get_ylim()[1] * 0.83, f'{label_INELAS}', color=color_INELAS, ha='center', va='bottom', fontsize=7, rotation=90)
            
            dm_INELASTIC_min_max_INELAS = {etiqueta_INELAS: [np.inf, -np.inf] for etiqueta_INELAS in etiquetas_INELAS}
            frecuencia_min_max_INELAS = {etiqueta_INELAS: [np.inf, -np.inf] for etiqueta_INELAS in etiquetas_INELAS}
            
            # Associate the minimum and maximum values ​​of dm_elastic and frequencies the inelastic dynamic magnification intervals
            for i_INELAS, (ELAS_SUBLIST_INELAS, frequ_INELASTIC_sublist_INELAS) in enumerate(zip(self.Rdnolineal_list[i], self.frecuencias_summary)):
                for j_INELAS, freq_INELAS in enumerate(ELAS_SUBLIST_INELAS):
                    for k_INELAS in range(self.num_divisiones_INELAS):
                        if intervalos_INELAS[k_INELAS] <= freq_INELAS < intervalos_INELAS[k_INELAS + 1]:
                            frecuencia_min_max_INELAS[etiquetas_INELAS[k_INELAS]][0] = min(frecuencia_min_max_INELAS[etiquetas_INELAS[k_INELAS]][0], freq_INELAS)
                            frecuencia_min_max_INELAS[etiquetas_INELAS[k_INELAS]][1] = max(frecuencia_min_max_INELAS[etiquetas_INELAS[k_INELAS]][1], freq_INELAS)
                            dm_INELASTIC_min_max_INELAS[etiquetas_INELAS[k_INELAS]][0] = min(dm_INELASTIC_min_max_INELAS[etiquetas_INELAS[k_INELAS]][0], frequ_INELASTIC_sublist_INELAS[j_INELAS])
                            dm_INELASTIC_min_max_INELAS[etiquetas_INELAS[k_INELAS]][1] = max(dm_INELASTIC_min_max_INELAS[etiquetas_INELAS[k_INELAS]][1], frequ_INELASTIC_sublist_INELAS[j_INELAS])
            
            # Create labels for minimum and maximum values ​​of dm_elastic and frequency associated with the inelastic dynamic magnification
            for etiqueta_INELAS in etiquetas_INELAS:
                if frecuencia_min_max_INELAS[etiqueta_INELAS][0] == np.inf:
                    frecuencia_min_max_INELAS[etiqueta_INELAS] = ['N/A', 'N/A']
                if dm_INELASTIC_min_max_INELAS[etiqueta_INELAS][0] == np.inf:
                    dm_INELASTIC_min_max_INELAS[etiqueta_INELAS] = ['N/A', 'N/A']
                if (frecuencia_min_max_INELAS[etiqueta_INELAS][0] == frecuencia_min_max_INELAS[etiqueta_INELAS][1] and
                    frecuencia_min_max_INELAS[etiqueta_INELAS][0] != 'N/A'):
                    frecuencia_min_max_INELAS[etiqueta_INELAS][1] = frecuencia_min_max_INELAS[etiqueta_INELAS][0]
                if (dm_INELASTIC_min_max_INELAS[etiqueta_INELAS][0] == dm_INELASTIC_min_max_INELAS[etiqueta_INELAS][1] and
                    dm_INELASTIC_min_max_INELAS[etiqueta_INELAS][0] != 'N/A'):
                    dm_INELASTIC_min_max_INELAS[etiqueta_INELAS][1] = dm_INELASTIC_min_max_INELAS[etiqueta_INELAS][0]
            # Create a group of labels first column inelastic dynamic magnification intervals second column minimum and maximum values ​​of dm_elastic
            # and finally third column minimum and maximum values ​​of frequencies
            leyenda_etiquetas_INELAS = []
            for etiqueta_INELAS in etiquetas_INELAS:
                if dm_INELASTIC_min_max_INELAS[etiqueta_INELAS][0] == 'N/A':
                    elastic_label_INELAS = 'N/A'
                else:
                    elastic_label_INELAS = f'{dm_INELASTIC_min_max_INELAS[etiqueta_INELAS][0]:.2f} - {dm_INELASTIC_min_max_INELAS[etiqueta_INELAS][1]:.2f}'
                
                leyenda_etiquetas_INELAS.append(
                    f'ψ_I {etiqueta_INELAS}; Freq. ({elastic_label_INELAS})Hz'
                )
            
            # Customizing the legends of the graph in which I add the values ​​of the 16th, 50th, 84th percentiles to the Gauss distribution
            # Shade the areas, show the probabilities associated with the already mentioned percentiles and also place the values ​​of frequencies and dynamic magnifications
            handles_INELAS = [
                plt.Line2D([0], [0], color=(0, 0, 0, 1), linestyle='-', linewidth=2, label='Gaussian Distribution of Inelastic Dynamic Magnification'),
                plt.Line2D([0], [0], color='SteelBlue', linestyle='--', linewidth=1, label=f'Percentile 16 = {percentiles_INELAS[0]:.2f}'),
                plt.Line2D([0], [0], color='DodgerBlue', linestyle='--', linewidth=1, label=f'Percentile 50 = {percentiles_INELAS[1]:.2f}'),
                plt.Line2D([0], [0], color='royalblue', linestyle='--', linewidth=1, label=f'Percentile 86 = {percentiles_INELAS[2]:.2f}'),
                plt.Line2D([0], [0], color=(0.27, 0.51, 0.71, 0.5), linestyle='-', linewidth=10, label=f'Pr(ψ_I≤p16) = {p_16_INELAS:.2f}%'),
                plt.Line2D([0], [0], color=(0.53, 0.81, 0.92, 0.5), linestyle='-', linewidth=10, label=f'Pr(p16<ψ_I≤p50) = {p_16_50_INELAS:.2f}%'),
                plt.Line2D([0], [0], color=(0.12, 0.56, 1, 0.5), linestyle='-', linewidth=10, label=f'Pr(p50<ψ_I≤p84) = {p_50_84_INELAS:.2f}%'),
                plt.Line2D([0], [0], color=(0.25, 0.41, 0.88, 0.5), linestyle='-', linewidth=10, label=f'Pr(ψ_I>p84) = {p_84_INELAS:.2f}%'),
                plt.Line2D([0], [0], color=(0.5, 0.5, 0.5, 0.5), linestyle='-', linewidth=10, label='I. Dynamic Magnification Histogram')
                #plt.Line2D([0], [0], color='none', linestyle='', linewidth=0, label='Inelastic Dynamic Magnification Intervals and Frequency:')
            ]
            
            #handles_INELAS.extend([
                #plt.Line2D([0], [0], color='white', linestyle='-', linewidth=0, marker='o', markersize=1, label=etiqueta_INELAS)
                #for etiqueta_INELAS in leyenda_etiquetas_INELAS
            #])
            
            # Addsymbols of the mean and standard deviation followed by their values ​​for the inelastic dynamic magnification
            mean_text_INELAS = f'$\overline{{x}} = {mean_INELAS:.2f} \ \mathrm{{Hz}}$'
            std_text_INELAS = f'σ = {std_INELAS:.2f} Hz'
            ax6.text(0.01, 0.98, mean_text_INELAS, transform=ax6.transAxes, fontsize=7.5, verticalalignment='top', horizontalalignment='left')
            ax6.text(0.01, 0.93, std_text_INELAS, transform=ax6.transAxes, fontsize=7.5, verticalalignment='top', horizontalalignment='left')
            
            # parameters to customize the graph adjust legends text sizes
            ax6.legend(handles=handles_INELAS, loc='upper left', bbox_to_anchor=(1, 1), title=None, frameon=False, fontsize=7.0) 
            ax6.yaxis.set_ticks([])
            plt.xticks(fontsize=7.5)  
            plt.yticks(fontsize=7.5)  
            plt.tight_layout()
            # Save plot summary Gaussian Distribution of Inelastic Dynamic Magnification
            nombre_archivos_graficas_MDIN_INELAS_distribuciong = f'summary_Gaussian_Distribution_of_Inelastic_Dynamic_Magnification_for_zeta_{self.amortiguamientos[i]*100:.2f}_porcentaje.png'
            self.localListImages.append(nombre_archivos_graficas_MDIN_INELAS_distribuciong)
            ruta_completa_imagenes_MDIN_INELAS_distribuciong = os.path.join(self.directorio_imagenes_general, nombre_archivos_graficas_MDIN_INELAS_distribuciong)
            plt.savefig(ruta_completa_imagenes_MDIN_INELAS_distribuciong, bbox_inches='tight')  # Guardar la figura en el directorio
            plt.close(fig)
            
            
            
        
        ########################################################################################################################
        # Report of results of both the frequency content of the seismic records and the calculation of elastic and inelastic dynamic magnification 
        # based on these 
        ########################################################################################################################

        num_freq_90_prs = []
        
        # Iterate over each sublist in self.frecuencias_summary and store its length as float
        for sublista in self.frecuencias_summary:
            num_freq_90_prs.append(float(len(sublista)))
    
        # Iterate over each sublist in self.energia_summary and store its length as float
        num_energia_90_prs = []
        for sublista in self.energia_summary:
            num_energia_90_prs.append(float(len(sublista)))
        
        # Define the output directory and file
        """ output_dir = r'D:/Temporal/MiguelRiveraCambios/temporalimagespruebas/alter/'
        output_file = "Summary Report.xlsx"
        output_path = os.path.join(output_dir, output_file) """
        output_path = self.datos_informe
        
        #os.makedirs(output_dir, exist_ok=True)
        
        # Create a new workbook and sheet to fill with the results of frequencies and dynamic magnifications
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Summary Report"
        
        # Define the bold style
        bold_font = Font(bold=True)
        
        # Add Titles to the first rows in summary reports
        sheet.cell(row=1, column=1, value="Damping").font = bold_font
        sheet.cell(row=1, column=2, value="ID").font = bold_font
        sheet.cell(row=1, column=3, value="Energy (%)").font = bold_font
        sheet.cell(row=1, column=4, value="Frequencies [Hz]").font = bold_font
        sheet.cell(row=1, column=5, value="ψ ELA. Dyn. Magn.").font = bold_font
        sheet.cell(row=1, column=6, value="ψ INE. Dyn. Magn.").font = bold_font  # Agregar encabezado para Rdnolineal con negritas
        
        # Set row and column values ​​for damping labels
        start_row_amortiguamiento = 2  
        start_col_amortiguamiento = 1  
        # Set row and column values ​​for id seism
        start_row_idsismo = 2  
        start_col_idsismo = 2 
        # Set row and column values ​​for energy
        start_row_energia = 2  
        start_col_energia = 3 
        # Set row and column values ​​for frequencies
        start_row_frecuencias = 2  
        start_col_frecuencias = 4  
        # Set row and column values ​​for Dynamic Magnification Elastic
        start_row_rdlineal = 2  
        start_col_rdlineal = 5  
        # Set row and column values ​​for Dynamic Magnification Inelastic
        start_row_rdnolineal = 2  
        start_col_rdnolineal = 6 
        
        
        def write_values(sheet, start_row, start_col, data, n_repeticiones=1):
            row = start_row
            for _ in range(n_repeticiones):
                for item in data:
                    if isinstance(item, list):  # Verificar si el item es una lista
                        for sublist in item:
                            if isinstance(sublist, list):  # Si el sublist es otra lista
                                for value in sublist:
                                    sheet.cell(row=row, column=start_col, value=value)
                                    row += 1
                            else:  # Si el sublist no es una lista, es un valor individual
                                sheet.cell(row=row, column=start_col, value=sublist)
                                row += 1
                    else:  # Si el item no es una lista
                        sheet.cell(row=row, column=start_col, value=item)
                        row += 1
            return row
        
        # Define the number of repetitions of some of the results.
        n_repeticiones = len(self.amortiguamientos)
        
        # Start writing energia_summary in the corresponding column
        write_values(sheet, start_row_energia, start_col_energia, self.energia_summary, n_repeticiones)
        
        # Start writing frequencias_planas in the corresponding column
        write_values(sheet, start_row_frecuencias, start_col_frecuencias, frecuencias_planas, n_repeticiones)
        
        # Start writing elastic dynamic magnification in the corresponding column
        write_values(sheet, start_row_rdlineal, start_col_rdlineal, self.Rdlineal_list)
        
        # Start writing inelastic dynamic magnification in the corresponding column
        write_values(sheet, start_row_rdnolineal, start_col_rdnolineal, self.Rdnolineal_list)
        
        # Merge cells vertically in ID column based on num_freq_90_prs values
        current_row_idsismo = start_row_idsismo 
        for _ in range(n_repeticiones): 
            for i, num_rows in enumerate(num_freq_90_prs):
                end_row_idsismo = current_row_idsismo + int(num_rows) - 1
                sheet.merge_cells(start_row=current_row_idsismo, start_column=start_col_idsismo, end_row=end_row_idsismo, end_column=start_col_idsismo)
                sheet.cell(row=current_row_idsismo, column=start_col_idsismo, value=f'S{i+1}')
                current_row_idsismo = end_row_idsismo + 1
        
        # Merge cells vertically in the buffer column based on the total length of self.frequencies_summary
        current_row_amortiguamiento = start_row_amortiguamiento
        for i in range(n_repeticiones):
            end_row_amortiguamiento = current_row_amortiguamiento + sum(num_freq_90_prs) - 1
            sheet.merge_cells(start_row=current_row_amortiguamiento, start_column=start_col_amortiguamiento, end_row=end_row_amortiguamiento, end_column=start_col_amortiguamiento)
            sheet.cell(row=current_row_amortiguamiento, column=start_col_amortiguamiento, value=f'ζ = {self.amortiguamientos[i] * 100:.2f}%')
            current_row_amortiguamiento = end_row_amortiguamiento + 1
        
        # Center the content of all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust the width of the columns to fit the content
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter 
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the Excel file of the frequency and dynamic magnification results report in the specified directory
        workbook.save(output_path)
        self.RenderFiles()
        return

    def RenderFiles(self):
        # Define a HTML tag for page break in rendered HTML output
        salto_linea = (
            '<div style = "display:block; clear:both; page-break-after:always;"></div>'
        )        
        # Append filename header to main HTML output
        #self.salidaHtml.append(
        #    f"<h3 style='font-size:18px'>Summary</h3><hr>"
        #)
        # Start a new div container for images in main HTML output
        self.salidaHtml.append('<div>')
        # Iterate through each image index
        for localImage in self.localListImages:
            #self.salidaHtml.append(f"<center><h3>{image_etiquetas[i-1]}</h3></center>") adfsdgfadsfadsfasdfasdfasdfadfasdfsadfadsf
            # Append image tag with source path and styling to main HTML output
            self.salidaHtml.append(
                f"<img src='images/summary/{localImage}' style=\"margin-bottom: 20px; border: 2px solid #000000; padding: 10px;\" width='98%'>" #cambio de 100 a 98 para indicar el borde
            )
        # Close the div container for images in main HTML output
        self.salidaHtml.append("</div>")
        # Combine all HTML elements into a single string for main HTML output
        self.htmlText = " ".join(self.salidaHtml)

        # Append filename header to temporary HTML output
        self.salidaTemporalHtml.append(
            f"<h3 style='font-size:18px'>Summary</h3><hr>"
        )
        iTratamientos = 1
        strTable = "<table border='1'><tr><th>Id</th><th>File</th></tr>"
        for tratamiento in self.tratamientos:
            strTable += f"<tr><td>S{iTratamientos}</td><td>{tratamiento.filebasename}</td></tr>"
            iTratamientos += 1
        strTable += "</table>"

        self.salidaTemporalHtml.append(strTable)
        self.salidaTemporalHtml.append(salto_linea)


        # Iterate through each image index for temporary HTML output
        i = 0
        for localImage in self.localListImages:

            # self.salidaTemporalHtml.append(f"<img src='file://{self.proyectPath}/results/html/images/{self.tratamiento.filebasename}/{i}.png' width='100%'><hr>")
            # Open the images of the summary of results of the seismic records considered in binary format to read them
            with open(
                f"{self.proyectPath}/results/html/images/summary/{localImage}",
                "rb",
            ) as image_file:
                # Read image data from file
                image_data = image_file.read()
            
            #if i != 1 :
            
            #self.salidaTemporalHtml.append(f"<center><h3>{image_etiquetas[i-1]}</h3></center>")   fgsfdgdsfgsfdgsdgfsdfgsdfgsdfgsdfgsdgfsfdg
            # Convert image data to Base64 format
            image_base64 = base64.b64encode(image_data).decode("utf-8")
            # Append image tag with Base64-encoded source to temporary HTML output
            self.salidaTemporalHtml.append(
                f'<img src="data:image/png;base64,{image_base64}" width="100%">'
            )
            # Add a page break after each image except the last one in temporary HTML output
            i = i + 1
            if(i < len(self.localListImages)-1):
                self.salidaTemporalHtml.append(salto_linea)
        # Combine all HTML elements into a single string for temporary HTML output
        self.temporalhtmlText = " ".join(self.salidaTemporalHtml)

        # Write the main HTML content to a text file
        with open(self.archivo_html, "w") as archivo:
            archivo.write(
                f"<html><head><title>Summary</title></head><body>"
            )
            archivo.write(self.htmlText)
            archivo.write("</body></html>")
        # Write the temporary HTML content to a separate text file
        with open(self.archivo_temporal_html, "w") as archivo:
            archivo.write(
                f"<html><head><title>Summary</title></head><body>"
            )
            archivo.write(self.temporalhtmlText)
            archivo.write("</body></html>")

        # Configure options for PDFKit
        opciones = {
            "page-size": "A4",
            "margin-top": "10mm",
            "margin-right": "10mm",
            "margin-bottom": "10mm",
            "margin-left": "10mm",
            "orientation": 'Landscape',
        }
        try:
            # Convert temporary HTML file to PDF using PDFKit
            pdfkit.from_file(
                self.archivo_temporal_html, self.archivo_pdf, options=opciones
            )
            print(f"Archivo PDF guardado en: {self.archivo_pdf}")
        except Exception as e:
            print(f"Error al convertir a PDF: {str(e)}")